import streamlit as st
import pandas as pd
from io import BytesIO
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib.utils import ImageReader
import streamlit.components.v1 as components

# ================= PAGE CONFIG (MUST BE FIRST STREAMLIT CALL) =================
st.set_page_config(page_title="Fixer-Upper Rental Analyzer", layout="wide")

# ================= TAB/TOP SPACING (KEEP YOUR STYLE) =================
st.markdown("""
<style>
.block-container { padding-top: 2rem !important; }
div[data-testid="stTabs"] { margin-top: 0 !important; padding-top: 0 !important; }
section.main > div { padding-top: 0 !important; }
</style>
""", unsafe_allow_html=True)

# ================= GOOGLE ANALYTICS =================
components.html(
    """
    <!-- Google tag (gtag.js) -->
    <script async src="https://www.googletagmanager.com/gtag/js?id=G-DMLKRR0K9P"></script>
    <script>
      window.dataLayer = window.dataLayer || [];
      function gtag(){dataLayer.push(arguments);}
      gtag('js', new Date());
      gtag('config', 'G-DMLKRR0K9P', { 'send_page_view': true });
    </script>
    """,
    height=0,
)

# ================= EXPORT HELPERS (USED BY RENTAL + FLIP) =================
LOGO_PATH = "assets/logo2.png"
APP_TITLE = "Rental Deal Analyzer"
APP_TAGLINE = "Know if a rental deal works — Fast & Free."

def _is_percent_field(key: str) -> bool:
    k = str(key).lower()
    return any(x in k for x in [
        "cap rate", "coc", "cash-on-cash", "equity", "vacancy", "management fee",
        "down payment", "cash % arv", "interest rate", "roi", "margin", "%", "commission",
        "annualized"
    ])

def _is_money_field(key: str) -> bool:
    k = str(key).lower()
    return any(x in k for x in [
        "rent", "noi", "cash flow", "debt", "down payment", "closing", "total",
        "expense", "tax", "insurance", "maintenance", "rehab", "loan", "investment",
        "price", "arv", "profit", "cost", "fees", "utilities", "hoa", "staging",
        "concessions", "permit", "points", "origination", "holding"
    ])

def _fmt_value(key, val):
    if isinstance(val, str):
        return val
    if isinstance(val, dict):
        return ""

    try:
        num = float(val)
    except Exception:
        return str(val)

    if _is_percent_field(str(key)):
        return f"{num * 100:.2f}%"
    if _is_money_field(str(key)):
        return f"${num:,.2f}"
    return f"{num:,.2f}"

def _build_export_rows(results: dict):
    rows = []
    for k, v in results.items():
        if isinstance(v, dict):
            rows.append((str(k), ""))
            for kk, vv in v.items():
                rows.append((f"  - {kk}", _fmt_value(kk, vv)))
        else:
            rows.append((str(k), _fmt_value(k, v)))
    return rows

def export_excel_pdf(results_dict: dict, excel_placeholder, pdf_placeholder, excel_filename: str, pdf_filename: str):
    export_rows = _build_export_rows(results_dict)
    export_df = pd.DataFrame(export_rows, columns=["Metric", "Value"])

    # -------- Excel (with logo + title) --------
    excel_buffer = BytesIO()
    try:
        with pd.ExcelWriter(excel_buffer, engine="openpyxl") as writer:
            export_df.to_excel(writer, index=False, sheet_name="Results", startrow=6)
            ws = writer.book["Results"]

            ws["A1"] = APP_TITLE
            ws["A2"] = APP_TAGLINE
            try:
                ws["A1"].font = ws["A1"].font.copy(bold=True, size=16)
                ws["A2"].font = ws["A2"].font.copy(size=11)
            except Exception:
                pass

            try:
                ws.merge_cells("A1:D1")
                ws.merge_cells("A2:D2")
            except Exception:
                pass

            try:
                from openpyxl.drawing.image import Image as XLImage
                logo = XLImage(LOGO_PATH)
                logo.width = 140
                logo.height = 40
                ws.add_image(logo, "E1")
            except Exception:
                pass

            try:
                ws.column_dimensions["A"].width = 34
                ws.column_dimensions["B"].width = 24
            except Exception:
                pass
    except Exception:
        export_df.to_excel(excel_buffer, index=False)

    excel_buffer.seek(0)
    excel_placeholder.download_button(
        "⬇️ Download Excel",
        excel_buffer,
        excel_filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    # -------- PDF (with logo + title) --------
    pdf_buffer = BytesIO()
    c = canvas.Canvas(pdf_buffer, pagesize=letter)

    top_y = 760
    try:
        logo_img = ImageReader(LOGO_PATH)
        c.drawImage(logo_img, 40, top_y - 35, width=120, height=35, mask="auto")
    except Exception:
        pass

    c.setFont("Helvetica-Bold", 16)
    c.drawString(170, top_y - 15, APP_TITLE)
    c.setFont("Helvetica", 10)
    c.drawString(170, top_y - 30, APP_TAGLINE)

    y = 700
    c.setFont("Helvetica", 10)

    for metric, value in export_rows:
        if value == "" and not str(metric).startswith("  - "):
            c.setFont("Helvetica-Bold", 11)
            c.drawString(40, y, str(metric))
            y -= 16
            c.setFont("Helvetica", 10)
        else:
            c.drawString(50, y, f"{metric}: {value}")
            y -= 14

        if y < 60:
            c.showPage()
            y = 750
            c.setFont("Helvetica-Bold", 12)
            c.drawString(40, y, APP_TITLE)
            y -= 20
            c.setFont("Helvetica", 10)

    c.save()
    pdf_buffer.seek(0)

    pdf_placeholder.download_button(
        "⬇️ Download PDF",
        pdf_buffer,
        pdf_filename,
        mime="application/pdf"
    )

# ================= TABS =================
tab_rental, tab_flip = st.tabs(["🏠 Rental Calculator", "🔨 Flip Calculator"])

# =====================================================================
# =========================== RENTAL TAB ===============================
# =====================================================================
with tab_rental:
    # ================= TOP BAR =================
    top_left, top_middle, top_right = st.columns([5, 2, 2])

    with top_left:
        st.markdown(
            """
            <div style="display:flex; align-items:center; gap:14px; margin:8px 0 8px;">
              <svg width="48" height="48" viewBox="0 0 64 64" xmlns="http://www.w3.org/2000/svg">
                <path d="M8 30L32 10L56 30V54H38V40H26V54H8V30Z"
                      fill="rgba(29,161,242,0.25)" stroke="#EAF0FF" stroke-width="2"/>
                <rect x="22" y="34" width="5" height="10" fill="#1DA1F2"/>
                <rect x="30" y="30" width="5" height="14" fill="#1DA1F2"/>
                <rect x="38" y="26" width="5" height="18" fill="#1DA1F2"/>
              </svg>
              <span style="font-size:30px; font-weight:800;">Rental Deal Analyzer</span>
            </div>
            """,
            unsafe_allow_html=True
        )

        st.markdown("""
### Know if a rental deal works — Fast & Free.  
Get **cash flow, cap rate, cash-on-cash return, deal score and much more** instantly.
""")

    with top_middle:
        breakdown_view = st.selectbox(
            "📊 View Mode",
            ["Annual", "Monthly"],
            index=0,
            key="breakdown_view_rental"
        )

    with top_right:
        excel_btn = st.empty()
        pdf_btn = st.empty()
        st.markdown(
            "<div style='font-size:14px; margin-top:4px;'>"
            "📧 Email: <a href='mailto:email@rentaldealanalyzer.com'>"
            "email@rentaldealanalyzer.com</a></div>",
            unsafe_allow_html=True
        )

    st.markdown(
        """
        🔒 **Privacy Notice:**  
        *We do not store or track your deal data. All calculations are performed in real-time and reset when you refresh the page.*
        """
    )

    col1, spacer1, col2, spacer2, col3 = st.columns([1.2, 0.15, 1, 0.15, 1])

    with col1:
        st.header("🔢 Deal Inputs")

        purchase_price = st.number_input("Purchase Price ($)", min_value=0, step=1000, value=0, key="purchase_price")
        rehab_cost = st.number_input("Rehab Cost ($)", min_value=0, step=1000, value=0, key="rehab_cost")
        arv = st.number_input("After Repair Value (ARV) ($)", min_value=0, step=1000, value=0, key="arv")

        monthly_rent = st.number_input("Monthly Rent ($)", min_value=0, step=100, value=0, key="monthly_rent")

        property_tax = st.number_input("Annual Property Tax ($)", min_value=0, step=100, value=0, key="property_tax")
        insurance = st.number_input("Annual Insurance ($)", min_value=0, step=100, value=0, key="insurance")
        maintenance = st.number_input("Annual Maintenance ($)", min_value=0, step=100, value=0, key="maintenance")

        vacancy_rate = st.number_input("Vacancy Rate (%)", 0, 100, value=0, key="vacancy_rate") / 100
        management_fee = st.number_input("Management Fee (%)", 0, 100, value=0, key="management_fee") / 100

        down_payment_pct = st.number_input("Down Payment (%)", 0, 100, value=0, key="down_payment_pct") / 100
        interest_rate = st.number_input(
            "Interest Rate (%)",
            min_value=0.0,
            max_value=15.0,
            value=0.0,
            step=0.25,
            format="%.2f",
            key="interest_rate_rental"
        ) / 100

        loan_term = st.number_input("Loan Term (Years)", 1, 40, value=30, key="loan_term")

        closing_cost_pct = st.number_input(
            "Estimated Closing Costs (% of Purchase Price)",
            min_value=0,
            max_value=10,
            value=3,
            key="closing_cost_pct"
        ) / 100

        analyze = st.button("📊 Analyze Deal", key="analyze_rental")

    if analyze:
        annual_rent = monthly_rent * 12
        vacancy_loss = annual_rent * vacancy_rate
        management_cost = annual_rent * management_fee

        expenses_annual = {
            "Property Tax": property_tax,
            "Insurance": insurance,
            "Maintenance": maintenance,
            "Vacancy Loss": vacancy_loss,
            "Management Fee": management_cost
        }

        total_expenses_annual = sum(expenses_annual.values())
        noi_annual = annual_rent - total_expenses_annual

        loan_amount = purchase_price * (1 - down_payment_pct)
        monthly_rate = interest_rate / 12
        total_payments = loan_term * 12

        if interest_rate > 0 and total_payments > 0:
            monthly_payment = loan_amount * (
                monthly_rate * (1 + monthly_rate) ** total_payments
            ) / ((1 + monthly_rate) ** total_payments - 1)
        else:
            monthly_payment = loan_amount / total_payments if total_payments else 0

        annual_debt = monthly_payment * 12
        cash_flow_annual = noi_annual - annual_debt

        total_investment = purchase_price + rehab_cost
        cash_invested = purchase_price * down_payment_pct + rehab_cost

        cap_rate = noi_annual / total_investment if total_investment else 0
        coc_return = cash_flow_annual / cash_invested if cash_invested else 0
        equity_pct = (arv - total_investment) / arv if arv else 0

        deal_score = min(
            100,
            (coc_return / 0.15 * 40) +
            (cap_rate / 0.10 * 30) +
            (equity_pct / 0.20 * 30)
        )

        rating = (
            "🔥 Excellent Deal" if deal_score >= 85 else
            "✅ Strong Deal" if deal_score >= 70 else
            "⚠️ Marginal Deal" if deal_score >= 50 else
            "❌ Weak Deal"
        )

        down_payment = purchase_price * down_payment_pct
        closing_costs = purchase_price * closing_cost_pct
        total_cash_needed = down_payment + rehab_cost + closing_costs
        cash_pct_arv = total_cash_needed / arv if arv else 0

        st.session_state.results = {
            "Annual Rent": annual_rent,
            "Monthly Rent": monthly_rent,
            "NOI Annual": noi_annual,
            "Cash Flow Annual": cash_flow_annual,
            "Debt Annual": annual_debt,
            "Debt Monthly": monthly_payment,
            "Cap Rate": cap_rate,
            "CoC": coc_return,
            "Equity": equity_pct,
            "Score": deal_score,
            "Rating": rating,
            "Expenses Annual": expenses_annual,
            "Total Expenses Annual": total_expenses_annual,
            "Down Payment": down_payment,
            "Closing Costs": closing_costs,
            "Total Cash Needed": total_cash_needed,
            "Cash % ARV": cash_pct_arv
        }

    with col2:
        st.header("📈 Deal Results")
        if "results" in st.session_state:
            r = st.session_state.results

            if breakdown_view == "Annual":
                st.metric("Rent", f"${r['Annual Rent']:,.0f}", help="Total gross rental income per year.")
                st.metric("NOI", f"${r['NOI Annual']:,.0f}", help="Net Operating Income before debt service.")
                st.metric("Cash Flow", f"${r['Cash Flow Annual']:,.0f}", help="Annual cash remaining after all expenses and debt service.")
                st.metric("Debt Service", f"${r['Debt Annual']:,.0f}", help="Total annual mortgage payments (principal + interest).")
            else:
                st.metric("Rent", f"${r['Monthly Rent']:,.0f}", help="Gross rental income per month.")
                st.metric("NOI", f"${r['NOI Annual']/12:,.0f}", help="Monthly Net Operating Income before debt service.")
                st.metric("Cash Flow", f"${r['Cash Flow Annual']/12:,.0f}", help="Monthly cash remaining after expenses and debt service.")
                st.metric("Debt Service", f"${r['Debt Monthly']:,.0f}", help="Monthly mortgage payment (principal + interest).")

            st.metric("Cap Rate", f"{r['Cap Rate']:.2%}", help="NOI divided by total investment cost.")
            st.metric("Cash-on-Cash Return", f"{r['CoC']:.2%}", help="Annual cash flow divided by cash invested.")
            st.metric("Equity %", f"{r['Equity']:.2%}", help="Percentage of property value owned after purchase and rehab.")
            st.metric("Rental Deal Score", f"{r['Score']:.0f}/100", help="Overall deal strength score based on returns and equity.")
            st.subheader(r["Rating"])

    with col3:
        st.header("💸 Expense Breakdown")
        if "results" in st.session_state:
            r = st.session_state.results

            if breakdown_view == "Annual":
                for name, value in r["Expenses Annual"].items():
                    st.write(f"**{name}**: ${value:,.0f}")
                st.write(f"**Total Expenses:** ${r['Total Expenses Annual']:,.0f}")
            else:
                for name, value in r["Expenses Annual"].items():
                    st.write(f"**{name}**: ${value/12:,.0f}")
                st.write(f"**Total Expenses:** ${r['Total Expenses Annual']/12:,.0f}")

            st.subheader("💰 Cash Required at Closing")
            st.write(f"Down Payment: ${r['Down Payment']:,.0f}")
            st.write(f"Rehab Budget: ${rehab_cost:,.0f}")
            st.write(f"Closing Costs: ${r['Closing Costs']:,.0f}")
            st.write(f"**Total Cash Needed:** ${r['Total Cash Needed']:,.0f}")
            st.write(f"Cash Needed as % of ARV: {r['Cash % ARV']:.1%}")

    if "results" in st.session_state:
        export_excel_pdf(
            st.session_state.results,
            excel_btn,
            pdf_btn,
            excel_filename="rental_deal_analysis.xlsx",
            pdf_filename="rental_deal_analysis.pdf"
        )

    st.markdown("---")
    st.caption(
        "Disclaimer: This tool is for educational and informational purposes only and does not "
        "constitute financial, investment, legal, tax, or real estate advice. "
        "All outputs are estimates, and use of this tool is at your own risk."
    )

# =====================================================================
# ============================ FLIP TAB ================================
# =====================================================================
with tab_flip:
    # ================= TOP BAR =================
    top_left_f, top_middle_f, top_right_f = st.columns([5, 2, 2])

    with top_left_f:
        st.markdown(
            """
            <div style="display:flex; align-items:center; gap:14px; margin:8px 0 8px;">
              <svg width="48" height="48" viewBox="0 0 64 64" xmlns="http://www.w3.org/2000/svg">
                <path d="M8 30L32 10L56 30V54H38V40H26V54H8V30Z"
                      fill="rgba(29,161,242,0.25)" stroke="#EAF0FF" stroke-width="2"/>
                <rect x="22" y="34" width="5" height="10" fill="#1DA1F2"/>
                <rect x="30" y="30" width="5" height="14" fill="#1DA1F2"/>
                <rect x="38" y="26" width="5" height="18" fill="#1DA1F2"/>
              </svg>
              <span style="font-size:30px; font-weight:800;">Fix & Flip Deal Analyzer</span>
            </div>
            """,
            unsafe_allow_html=True
        )

        st.markdown("""
### Know if a flip works — Fast & Free.  
Get **profit, ROI, annualized ROI, break-even price, and total project cost** instantly.
""")

    with top_middle_f:
        st.selectbox("📊 View Mode", ["Project"], index=0, key="flip_view_mode", disabled=True)

    with top_right_f:
        flip_excel_btn = st.empty()
        flip_pdf_btn = st.empty()
        st.markdown(
            "<div style='font-size:14px; margin-top:4px;'>"
            "📧 Email: <a href='mailto:email@rentaldealanalyzer.com'>"
            "email@rentaldealanalyzer.com</a></div>",
            unsafe_allow_html=True
        )

    st.markdown(
        """
        🔒 **Privacy Notice:**  
        *We do not store or track your deal data. All calculations are performed in real-time and reset when you refresh the page.*
        """
    )

    fcol1, fsp1, fcol2, fsp2, fcol3 = st.columns([1.2, 0.15, 1, 0.15, 1])

    # Ensure session flags exist
    if "flip_analyzed" not in st.session_state:
        st.session_state.flip_analyzed = False

    # ================= LEFT COLUMN — FLIP INPUTS (ALL MISSING ADDED) =================
    with fcol1:
        st.header("🔢 Flip Inputs")

        st.subheader("Acquisition & Purchase")
        flip_purchase_price = st.number_input("Purchase Price ($)", min_value=0, step=1000, value=0, key="flip_purchase_price")

        # Buying Closing Costs: $ and/or %
        flip_buy_close_pct = st.number_input("Buying Closing Costs (% of Purchase Price)", min_value=0.0, max_value=15.0, value=2.50, step=0.25, format="%.2f", key="flip_buy_close_pct") / 100
        flip_buy_close_fixed = st.number_input("Buying Closing Costs ($)", min_value=0, step=100, value=0, key="flip_buy_close_fixed")

        flip_inspection_appraisal = st.number_input("Inspection & Appraisal Fees ($)", min_value=0, step=50, value=0, key="flip_inspection_appraisal")

        st.subheader("Renovation & Rehab")
        flip_rehab_budget = st.number_input("Estimated Rehab Budget ($)", min_value=0, step=1000, value=0, key="flip_rehab_budget")
        flip_contingency_pct = st.number_input("Misc / Contingency (%)", min_value=0.0, max_value=30.0, value=10.0, step=1.0, format="%.1f", key="flip_contingency_pct") / 100
        flip_permit_arch_fees = st.number_input("Permit & Architectural Fees ($)", min_value=0, step=100, value=0, key="flip_permit_arch_fees")

        st.subheader("Carrying Costs (Monthly)")
        flip_tax_mo = st.number_input("Property Taxes ($/month)", min_value=0, step=50, value=0, key="flip_tax_mo")
        flip_ins_mo = st.number_input("Insurance ($/month)", min_value=0, step=50, value=0, key="flip_ins_mo")
        flip_utils_mo = st.number_input("Utilities ($/month)", min_value=0, step=50, value=0, key="flip_utils_mo")
        flip_hoa_mo = st.number_input("HOA Fees ($/month)", min_value=0, step=25, value=0, key="flip_hoa_mo")
        flip_yard_pool_mo = st.number_input("Yard/Pool Maintenance ($/month)", min_value=0, step=25, value=0, key="flip_yard_pool_mo")

        st.subheader("Financing (If using a loan)")
        flip_down_payment_pct = st.number_input("Down Payment (%)", min_value=0.0, max_value=100.0, value=20.0, step=1.0, format="%.1f", key="flip_down_payment_pct") / 100
        flip_interest_rate = st.number_input("Interest Rate (Annual %)", min_value=0.0, max_value=25.0, value=12.0, step=0.25, format="%.2f", key="flip_interest_rate") / 100

        # Points: $ and/or %
        flip_points_pct = st.number_input("Loan Points / Origination Fees (% of Loan)", min_value=0.0, max_value=10.0, value=2.0, step=0.25, format="%.2f", key="flip_points_pct") / 100
        flip_points_fixed = st.number_input("Loan Points / Origination Fees ($)", min_value=0, step=100, value=0, key="flip_points_fixed")

        flip_holding_months = st.number_input("Projected Holding Period (Months)", min_value=1, step=1, value=6, key="flip_holding_months")

        st.subheader("Sale & Exit")
        flip_arv = st.number_input("After Repair Value (ARV) ($)", min_value=0, step=1000, value=0, key="flip_arv")
        flip_sell_cost_pct = st.number_input("Selling Costs / Realtor Commission (% of Sale Price)", min_value=0.0, max_value=15.0, value=8.0, step=0.25, format="%.2f", key="flip_sell_cost_pct") / 100
        flip_seller_concessions = st.number_input("Seller Concessions ($)", min_value=0, step=100, value=0, key="flip_seller_concessions")
        flip_staging_photo = st.number_input("Staging & Photography ($)", min_value=0, step=100, value=0, key="flip_staging_photo")

        analyze_flip = st.button("📊 Analyze Flip", key="analyze_flip")

    # ================= FLIP CALCULATIONS (SAFE + COMPLETE) =================
    if analyze_flip:
        # 1) Acquisition costs
        buying_closing_costs = (flip_purchase_price * flip_buy_close_pct) + flip_buy_close_fixed
        contingency_cost = flip_rehab_budget * flip_contingency_pct

        # 2) Carrying costs
        monthly_carry = flip_tax_mo + flip_ins_mo + flip_utils_mo + flip_hoa_mo + flip_yard_pool_mo
        total_carry = monthly_carry * flip_holding_months

        # 3) Financing
        down_payment_amt = flip_purchase_price * flip_down_payment_pct
        loan_amount = max(0.0, flip_purchase_price - down_payment_amt)

        # Many flip loans are interest-only → model interest-only carrying cost
        monthly_interest_payment = loan_amount * (flip_interest_rate / 12) if loan_amount > 0 else 0.0
        total_interest_paid = monthly_interest_payment * flip_holding_months

        points_cost = (loan_amount * flip_points_pct) + flip_points_fixed

        # 4) Selling costs (percent + fixed)
        selling_commission = flip_arv * flip_sell_cost_pct
        selling_fixed = flip_seller_concessions + flip_staging_photo

        # 5) Total project cost
        total_project_cost = (
            flip_purchase_price
            + buying_closing_costs
            + flip_inspection_appraisal
            + flip_rehab_budget
            + contingency_cost
            + flip_permit_arch_fees
            + total_carry
            + total_interest_paid
            + points_cost
            + selling_commission
            + selling_fixed
        )

        # 6) Profit
        net_profit = flip_arv - total_project_cost

        # Cash invested = total project cost minus borrowed principal (loan_amount)
        cash_invested = total_project_cost - loan_amount
        roi = (net_profit / cash_invested) if cash_invested else 0.0

        # Annualized ROI (compounded) – safe guards
        if flip_holding_months and roi > -1:
            annualized_roi = (1 + roi) ** (12 / flip_holding_months) - 1
        else:
            annualized_roi = 0.0

        # 7) Break-even sale price (commission depends on sale price)
        # Break-even S such that S - base_costs - (commission_pct*S) = 0
        # base_costs excludes commission computed on ARV and replaces with commission_pct*S
        base_costs = total_project_cost - selling_commission
        if flip_sell_cost_pct < 1:
            break_even_sale_price = base_costs / (1 - flip_sell_cost_pct)
        else:
            break_even_sale_price = 0.0

        # Store results + analyzed flag
        st.session_state.flip_analyzed = True
        st.session_state.flip_results = {
            "Purchase Price": flip_purchase_price,
            "Buying Closing Costs": buying_closing_costs,
            "Inspection & Appraisal Fees": flip_inspection_appraisal,
            "Estimated Rehab Budget": flip_rehab_budget,
            "Contingency Cost": contingency_cost,
            "Permit & Architectural Fees": flip_permit_arch_fees,
            "Monthly Carrying Costs": monthly_carry,
            "Total Carrying Costs": total_carry,
            "Down Payment": down_payment_amt,
            "Loan Amount": loan_amount,
            "Interest Rate": flip_interest_rate,
            "Monthly Interest Payment": monthly_interest_payment,
            "Total Interest Paid": total_interest_paid,
            "Loan Points / Origination Fees": points_cost,
            "After Repair Value (ARV)": flip_arv,
            "Selling Commission": selling_commission,
            "Seller Concessions": flip_seller_concessions,
            "Staging & Photography": flip_staging_photo,
            "Total Project Cost": total_project_cost,
            "Cash Invested": cash_invested,
            "Net Profit": net_profit,
            "ROI": roi,
            "Annualized ROI": annualized_roi,
            "Break-Even Sale Price": break_even_sale_price,
            "Holding Period (Months)": flip_holding_months,
        }

        st.session_state.flip_breakdown = {
            "Acquisition & Purchase": {
                "Purchase Price": flip_purchase_price,
                "Buying Closing Costs": buying_closing_costs,
                "Inspection & Appraisal Fees": flip_inspection_appraisal,
            },
            "Renovation & Rehab": {
                "Estimated Rehab Budget": flip_rehab_budget,
                "Contingency Cost": contingency_cost,
                "Permit & Architectural Fees": flip_permit_arch_fees,
            },
            "Carrying Costs": {
                "Monthly Carrying Costs": monthly_carry,
                "Total Carrying Costs": total_carry,
                "Monthly Interest Payment": monthly_interest_payment,
                "Total Interest Paid": total_interest_paid,
            },
            "Sale & Exit": {
                "Selling Commission": selling_commission,
                "Seller Concessions": flip_seller_concessions,
                "Staging & Photography": flip_staging_photo,
            }
        }

    # ================= MIDDLE COLUMN — FLIP RESULTS (DO NOT DISAPPEAR) =================
# ================= MIDDLE COLUMN — FLIP RESULTS =================
 with fcol2:
    st.header("📈 Flip Results")

    if "flip_results" in st.session_state:
        r = st.session_state.flip_results

        st.metric(
            "Net Profit",
            f"${r['Net Profit']:,.0f}",
            help="Final profit after all purchase, rehab, holding, and selling costs."
        )

        st.metric(
            "ROI",
            f"{r['ROI']:.2%}",
            help="Return on Investment: profit divided by total project cost."
        )

        st.metric(
            "Annualized ROI",
            f"{r['Annualized ROI']:.2%}",
            help="ROI adjusted to a 12-month holding period."
        )

        st.metric(
            "Profit Margin",
            f"{r['Profit Margin']:.2%}",
            help="Profit as a percentage of the sale price (ARV)."
        )

        st.metric(
            "Flip Deal Score",
            f"{r['Flip Deal Score']:.0f}/100",
            help="Overall flip quality score based on ROI, margin, and risk factors."
        )

        st.subheader(
            r["Rating"],
            help="Quick qualitative assessment of flip risk and return."
        )
    else:
        st.info("Enter inputs and click **Analyze Flip**")


    # ================= RIGHT COLUMN — COST BREAKDOWN (DO NOT DISAPPEAR) =================
    with fcol3:
        st.header("💸 Cost Breakdown")

        if st.session_state.get("flip_analyzed") and "flip_results" in st.session_state:
            r = st.session_state.flip_results

            st.write(f"Purchase Price: ${r['Purchase Price']:,.0f}")
            st.write(f"Buying Closing Costs: ${r['Buying Closing Costs']:,.0f}")
            st.write(f"Inspection & Appraisal: ${r['Inspection & Appraisal Fees']:,.0f}")
            st.write(f"Rehab Budget: ${r['Estimated Rehab Budget']:,.0f}")
            st.write(f"Contingency Cost: ${r['Contingency Cost']:,.0f}")
            st.write(f"Permits/Architect: ${r['Permit & Architectural Fees']:,.0f}")
            st.write(f"Total Carrying Costs: ${r['Total Carrying Costs']:,.0f}")
            st.write(f"Total Interest Paid: ${r['Total Interest Paid']:,.0f}")
            st.write(f"Loan Points/Fees: ${r['Loan Points / Origination Fees']:,.0f}")
            st.write(f"Selling Commission: ${r['Selling Commission']:,.0f}")
            st.write(f"Seller Concessions: ${r['Seller Concessions']:,.0f}")
            st.write(f"Staging & Photo: ${r['Staging & Photography']:,.0f}")
            st.markdown("---")
            st.write(f"**Total Project Cost:** ${r['Total Project Cost']:,.0f}")
            st.write(f"**ARV:** ${r['After Repair Value (ARV)']:,.0f}")

    # ================= FLIP DOWNLOAD BUTTONS (SEPARATE + DO NOT RESET VIEW) =================
    if st.session_state.get("flip_analyzed") and "flip_results" in st.session_state:
        export_excel_pdf(
            {
                **st.session_state.flip_results,
                "Cost Breakdown": st.session_state.get("flip_breakdown", {})
            },
            flip_excel_btn,
            flip_pdf_btn,
            excel_filename="flip_deal_analysis.xlsx",
            pdf_filename="flip_deal_analysis.pdf"
        )

    st.markdown("---")
    st.caption(
        "Disclaimer: This tool is for educational and informational purposes only and does not "
        "constitute financial, investment, legal, tax, or real estate advice. "
        "All outputs are estimates, and use of this tool is at your own risk."
    )
